from openpyxl import utils
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill

import os

# Define cell color
red = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

green = PatternFill(start_color='FF00FF00',
                   end_color='FF00FF00',
                   fill_type='solid')

yellow = PatternFill(start_color='00FFFF00',
                   end_color='00FFFF00',
                   fill_type='solid')

passed = 'Passed'

def get_workbook(wb_name):
    if os.path.isfile(wb_name):
        wb = load_workbook(wb_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Recognition Rate'
        ws['A1'] = 'File Name'
        # Set column width
        ws.column_dimensions[utils.get_column_letter(1)].width = 25
        ws['B1'] = 'Expected Results'
        ws.column_dimensions[utils.get_column_letter(2)].width = 20
        ws['C1'] = 'ZBar'
        ws.column_dimensions[utils.get_column_letter(3)].width = 20
        ws['D1'] = 'DBR'
        ws.column_dimensions[utils.get_column_letter(4)].width = 20
        ws['E1'] = 'ZXing'
        ws.column_dimensions[utils.get_column_letter(5)].width = 20
    return wb

def save_workbook(wb, wb_name):
    if wb != None:
        wb.save(wb_name)

def append_row(wb, filename=None, expected_results=None, zbar_results=None, dbr_results=None, ZXing_results=None):
    ws = wb.active
    ws.append([filename, expected_results, zbar_results, dbr_results, ZXing_results])

def update_row(wb, row_index, filename=None, expected_results=None, zbar_results=None, dbr_results=None, ZXing_results=None):
    ws = wb.active
    row = ws[row_index]
    row[0].value = filename
    row[1].value = expected_results
    if zbar_results != None:
        row[2].value = zbar_results
        if zbar_results == expected_results:
            row[2].fill = green
        else:
            row[2].fill = red

    if dbr_results != None:
        row[3].value = dbr_results
        if dbr_results == expected_results:
            row[3].fill = green
        else:
            row[3].fill = red

    if ZXing_results != None:
        row[4].value = ZXing_results
        if ZXing_results == expected_results:
            row[4].fill = green
        else:
            row[4].fill = red

def set_recognition_rate(wb, row_index, r1=None, r2=None, r3=None):
    ws = wb.active
    row = ws[row_index]
    row[2].value = r1
    row[3].value = r2
    row[4].value = r3
        

# Test
# name = 'data.xlsx'
# wb = get_workbook(name)
# ws = wb.active
# index = 2
# update_row(wb, index, r'D:\python-zxing-zbar-dbr\dataset\20499525_2.jpg', '20499525', '20499525', '20499525', '20499521')
# index += 1
# set_recognition_rate(wb, index, '59.46%', '75.68%', '13.51%')
# save_workbook(wb, name)




